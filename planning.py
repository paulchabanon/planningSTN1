import sys
import time
import json
import openpyxl
import ConfigParser
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from config import PlanningConfig
from gcalendar import Gcalendar
from shifts import Shifts

print '### START ###'

print 'open config'
config = PlanningConfig()

print 'connect to calendar'
gcal = Gcalendar(config.calendarName,config.calendarGMT,config.api['clientid'],config.api['secret'])
gcal.connect()

shifts = Shifts(gcal,config)

print 'open excel file'
wb = load_workbook(config.path)
ws = wb[config.tab_name]

if config.parse_days < 1:
  config.parse_days = float("inf")

user_col = -1
# search for trigram
for c in range(1,40):
  if ws.cell(row = config.trigram_row, column = c).value == config.trigram:
    user_col = c
    
if user_col < 0:
  print 'could not find trigram ', config.trigram
  sys.exit(1)

print 'found user',"\n"

#go for each day of year
count = 0
for r in range(1,380):
  day  = ws.cell(row = r, column = config.date_col).value
  if not type(day) is datetime:
    continue
  
  # get shift name (matin, soiree, etc...)
  shift = ws.cell(row = r, column = user_col).value
  if not type(shift) is unicode: # excel file is 100% unicode
    continue
  
  # this day is in the past, skip it
  if day < datetime.today():
    continue
  
  # max days to add in calendar reached
  if count >= config.parse_days:
    print 'max count reached, stopping'
    break
  
  count += 1
  
  # set date to the right format
  str_day = day.strftime('%Y-%m-%d')
  if str_day >= config.max_date:
    print 'max date reached, stopping'
    break
      
  #if shift name is not known in config file
  if shift in config.day_type.viewkeys():
    print str_day,config.day_type[shift]['start'],config.day_type[shift]['end'],'count',count
    if shifts.setShift(shift, str_day):
      time.sleep(1)
    
  else:
    print str_day,shift,'count',count
    shifts.rmShift(str_day)
    time.sleep(1)
  
  print "\n"
  
print '### END ###'
